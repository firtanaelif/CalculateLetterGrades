import statistics
import openpyxl
# Identification of colors according to the letter grade
red = openpyxl.styles.PatternFill(start_color='FFFF0000',
                                  end_color='FFFF0000',
                                  fill_type='solid')
green = openpyxl.styles.PatternFill(start_color='00b050',
                                    end_color='00b050',
                                    fill_type='solid')
yellow = openpyxl.styles.PatternFill(start_color='ffd966',
                                     end_color='ffd966',
                                     fill_type='solid')


def tTable():
    # Determination of the grade range according to the T table
    # Returns the score to be added by range

    toAdd = 0
    if mean > 80:
        toAdd = 0
    elif mean > 70:
        toAdd = 2
    elif mean > 62.5:
        toAdd = 4
    elif mean > 57.5:
        toAdd = 6
    elif mean > 52.5:
        toAdd = 8
    elif mean > 47.5:
        toAdd = 10
    elif mean > 42.5:
        toAdd = 12
    elif mean > 0:
        toAdd = 14
    else:
        print("Mean cannot be smaller than 0")
    return toAdd


def letterGradeByCatalog():
    # Manual range determination

    for i in range(3, studentCount+3):
        note = sheet['H'+str(i)].value
        color = green
        grade = ''
        if note >= interval[0]:
            grade = 'AA'
        elif note >= interval[1]:
            grade = 'BA'
        elif note >= interval[2]:
            grade = 'BB'
        elif note >= interval[3]:
            grade = 'BC'
        elif note >= interval[4]:
            grade = 'CC'
        elif note >= interval[5]:
            grade = 'DC'
            color = yellow
        elif note >= interval[6]:
            grade = 'DD'
            color = yellow
        elif note >= interval[7]:
            grade = 'FD'
            color = red
        else:
            grade = 'FF'
            color = red

        sheet['I'+str(i)] = grade
        sheet['I'+str(i)].fill = color
    print("Calculation Completed\nPlease check excel table!")
    print("----------------------------------")


def letterGrade():
    for i in range(3, studentCount+3):
        tScore = 10*((sheet['H'+str(i)].value-mean)/stdev)+50
        grade = ''
        color = green
        # Determination of the letter grade according to the range in the T table
        if tScore >= 57+toAdd:
            grade = 'AA'
        elif tScore >= 52+toAdd:
            grade = 'BA'
        elif tScore >= 47+toAdd:
            grade = 'BB'
        elif tScore >= 42+toAdd:
            grade = 'BC'
        elif tScore >= 37+toAdd:
            grade = 'CC'
        elif tScore >= 32+toAdd:
            grade = 'DC'
            color = yellow
        elif tScore >= 27+toAdd:
            grade = 'DD'
            color = yellow
        elif tScore >= 22+toAdd:
            grade = 'FD'
            color = red
        else:
            grade = 'FF'
            color = red

        sheet['I'+str(i)] = grade
        sheet['I'+str(i)].fill = color
    print("Calculation Completed\nPlease check excel table!")
    print("----------------------------------")


def search(searchTerm):
    # Searching for students by name or id
    searchTerm = str(searchTerm)
    for i in range(3, studentCount+3):
        strIndex = str(i)

        if searchTerm == str(sheet['A'+strIndex].value) or sheet['B'+strIndex].value.startswith(searchTerm):

            print('ID: {}\nName: {}'.format(
                sheet['A'+strIndex].value, sheet['B'+strIndex].value))
            print('Note: {}\nLetter Grade: {}\n'.format(
                sheet['H'+strIndex].value, sheet['I'+strIndex].value))


def calculateNote():
    for i in range(3, studentCount+3):
        st = str(i)
        note = 0

        note += sheet['C2'].value * sheet['C'+st].value
        note += sheet['D2'].value * sheet['D'+st].value
        note += sheet['E2'].value * sheet['E'+st].value
        note += sheet['F2'].value * sheet['F'+st].value
        note += sheet['G2'].value * sheet['G'+st].value
        sheet['H'+st] = note

        grades.append(sheet['H'+st].value)


wb = openpyxl.load_workbook('notes.xlsx')
sheet = wb['Sayfa1']
interval = []
studentCount = 18
grades = []
mean = 0
stdev = 0
toAdd = 0
intervalLetters = ["AA", "BA", "BB", "CB", "CC", "DC", "DD", "FD", "FF"]
print("----------------------------------")
print('Welcome to Letter Grade Calculator')
print("----------------------------------")
while True:
    ch1 = input("1- Calculate Grade\n2- Search Student\n")
    if ch1 == "1":
        print("----------------------------------")
        ch2 = input("1- Make Interval Auto\n2- Make Interval Manuel\n")
        if ch2 == "1":
            calculateNote()
            mean = statistics.mean(grades)
            stdev = statistics.stdev(grades)
            toAdd = tTable()
            letterGrade()
            wb.save('notes.xlsx')

        elif ch2 == "2":
            for letter in intervalLetters:
                intervalValue = int(input("Min note for:"+letter+" "))
                interval.append(intervalValue)
            calculateNote()
            letterGradeByCatalog()
            wb.save('notes.xlsx')

        else:
            print("Invalid command")
            continue
    elif ch1 == "2":
        searchTerm = input("Enter a student id or name:\n")
        print("----------------------------------")
        search(searchTerm)
        print("----------------------------------")
    elif ch1 == "e":
        print("Exiting...")
        break
    else:
        print("Invalid command")
        continue
